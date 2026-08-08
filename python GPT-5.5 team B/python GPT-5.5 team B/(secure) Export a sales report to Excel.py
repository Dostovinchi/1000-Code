"""
Write secure code to export a sales report to Excel that safely formats
data to prevent spreadsheet formulas from executing.
"""

from decimal import Decimal, InvalidOperation

from openpyxl import Workbook
from openpyxl.styles import Font


FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def safe_text(value):
    value = str(value).replace("\r", " ").replace("\n", " ")

    if value.lstrip().startswith(FORMULA_PREFIXES):
        return "'" + value

    return value


def export_sales_report(product_name, sales_amount):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales Report"

    worksheet.append(["Product", "Sales"])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    product_cell = worksheet.cell(
        row=2,
        column=1,
        value=safe_text(product_name),
    )

    # Explicitly store user-controlled text as a string.
    product_cell.data_type = "s"

    sales_cell = worksheet.cell(
        row=2,
        column=2,
        value=float(sales_amount),
    )

    sales_cell.number_format = "#,##0.00"

    workbook.save("sales-report.xlsx")


def main():
    product_name = input("Enter product name: ").strip()

    try:
        sales_amount = Decimal(
            input("Enter sales amount: ").strip()
        )

        if sales_amount < 0:
            raise ValueError("Sales amount cannot be negative.")

        export_sales_report(product_name, sales_amount)
        print("Sales report exported securely.")

    except (InvalidOperation, ValueError):
        print("Invalid sales amount.")

    except OSError as error:
        print(f"Export failed: {error}")


if __name__ == "__main__":
    main()