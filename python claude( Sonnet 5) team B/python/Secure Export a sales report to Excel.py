# Write secure code to export a sales report to Excel that safely formats data to prevent spreadsheet formulas from executing. 

import openpyxl
from openpyxl.styles import numbers

# Characters that trigger formula/DDE interpretation in Excel
DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def sanitize_value(value):
    """
    Neutralize any value that could be interpreted as a formula.
    Prefixing with a single quote forces Excel/openpyxl to treat
    the content as literal text rather than executing it.
    """
    if isinstance(value, str) and value.strip() and value.strip()[0] in DANGEROUS_PREFIXES:
        return "'" + value
    return value


def export_sales_report_secure(sales_data, output_path="sales_report.xlsx"):
    """
    Securely export sales data to an Excel file.
    All user-supplied fields are sanitized against formula injection
    and written using the 'text' number format as defense in depth.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ["Customer Name", "Product", "Quantity", "Notes", "Email"]
    ws.append(headers)

    for row in sales_data:
        customer_name = sanitize_value(row.get("customer_name"))
        product = sanitize_value(row.get("product"))
        quantity = row.get("quantity")  # numeric, no sanitization needed
        notes = sanitize_value(row.get("notes"))
        email = sanitize_value(row.get("email"))

        ws.append([customer_name, product, quantity, notes, email])

        # Defense in depth: force text-only cells for free-form string
        # fields, so even if sanitization is bypassed, Excel won't
        # evaluate the content as a formula.
        current_row = ws.max_row
        for col in ("A", "B", "D", "E"):  # skip Quantity (numeric column C)
            ws[f"{col}{current_row}"].number_format = numbers.FORMAT_TEXT

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    sample_data = [
        {
            "customer_name": "Jane Doe",
            "product": "Widget",
            "quantity": 5,
            "notes": "=HYPERLINK(\"http://evil.com\",\"Click\")",  # neutralized
            "email": "jane@example.com",
        },
        {
            "customer_name": "+cmd|'/c calc'!A0",  # neutralized
            "product": "Gadget",
            "quantity": 2,
            "notes": "Standard delivery",
            "email": "buyer@example.com",
        },
    ]
    export_sales_report_secure(sample_data)